import calendar

import openpyxl
from django.contrib import messages
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from main.admin import admin_site
from main.models import Absensi, Kelas, Siswa


def import_siswa(request):
    if request.method == "POST" and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        
        # Validasi ekstensi file
        if not excel_file.name.endswith('.xlsx'):
            messages.error(request, 'Format file harus .xlsx')
            return redirect('admin:import_siswa')

        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            
            success_count = 0
            skipped_count = 0
            
            # Loop mulai dari baris ke-2 (asumsi baris 1 adalah Header)
            for row in ws.iter_rows(min_row=2, values_only=True):
                # Ambil data berdasarkan urutan kolom: Nama, Kelas, NIS, NISN
                # row[0] = Nama, row[1] = Kelas, row[2] = NIS, row[3] = NISN
                
                raw_nama = row[0]
                raw_kelas = row[1]
                raw_nis = row[2]
                raw_nisn = row[3]

                # 1. Validasi Data Wajib (Nama & Kelas tidak boleh kosong)
                if not raw_nama or not raw_kelas:
                    continue # Skip baris kosong/rusak

                # Konversi ke string dan bersihkan spasi
                nama = str(raw_nama).strip()
                kelas_name = str(raw_kelas).strip()
                
                # Handle NIS & NISN (Bisa None/Kosong)
                # Kita pastikan kalau kosong jadi None, kalau ada jadi String
                nis = str(raw_nis).strip() if raw_nis else None
                nisn = str(raw_nisn).strip() if raw_nisn else None

                # 2. Cek Duplikasi (Skip logic)
                # Jika NIS ada isinya DAN sudah terdaftar -> Skip
                if nis and Siswa.objects.filter(nis=nis).exists():
                    skipped_count += 1
                    continue
                
                # Jika NISN ada isinya DAN sudah terdaftar -> Skip
                if nisn and Siswa.objects.filter(nisn=nisn).exists():
                    skipped_count += 1
                    continue

                # 3. Logic Kelas (Get or Create)
                # defaults={'active': True} artinya jika buat baru, set active=True
                kelas_obj, _ = Kelas.objects.get_or_create(
                    name=kelas_name,
                    defaults={'active': True} 
                )

                # 4. Simpan Siswa
                Siswa.objects.create(
                    fullname=nama,
                    kelas=kelas_obj,
                    nis=nis,
                    nisn=nisn
                )
                success_count += 1

            messages.success(request, f'Import Selesai! Berhasil: {success_count}, Skipped (Duplikat): {skipped_count}')
            return redirect('admin:import_siswa')

        except Exception as e:
            messages.error(request, f'Terjadi kesalahan saat memproses file: {str(e)}')
            return redirect('admin:import_siswa')

    context = admin_site.each_context(request)
    context['title'] = 'Import Siswa'
    return render(request, 'admin/import_siswa.html', context)


def export_absensi(request):
    # Jika method GET, tampilkan form filter
    if request.method == 'GET':
        context = {
            'kelas_list': Kelas.objects.filter_domain(request).filter(active=True),
            'years': range(2020, timezone.now().year + 2), # Rentang tahun
            'months': range(1, 13),
            'title': 'Export Absensi',
            **admin_site.each_context(request),
        }
        return render(request, 'admin/export_absensi.html', context)

    # Jika method POST, proses export
    elif request.method == 'POST':
        try:
            kelas_id = request.POST.get('kelas')
            month = int(request.POST.get('month'))
            year = int(request.POST.get('year'))
            
            kelas = get_object_or_404(Kelas, id=kelas_id)
            
            # Setup Workbook Excel
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            filename = f"Absensi_{kelas.name}_{year}_{month}.xlsx"
            response['Content-Disposition'] = f'attachment; filename={filename}'

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Rekap Absensi"

            # --- Styling ---
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            center_align = Alignment(horizontal='center', vertical='center')

            # --- Header Baris 1: Judul ---
            ws.merge_cells('A1:AK1') # Merge sampai kolom total
            ws['A1'] = f"LAPORAN ABSENSI KELAS {kelas.name} - PERIODE {month}/{year}"
            ws['A1'].font = Font(size=14, bold=True)
            ws['A1'].alignment = center_align

            # --- Header Baris 2: Kolom ---
            headers = ['No', 'NIS', 'Nama Siswa']
            
            # Generate angka tanggal sesuai jumlah hari di bulan tersebut
            _, num_days = calendar.monthrange(year, month)
            date_columns = [str(d) for d in range(1, num_days + 1)]
            
            summary_headers = ['H', 'S', 'I', 'A', 'B'] # Total summary
            
            all_headers = headers + date_columns + summary_headers

            # Tulis Header ke Excel
            for col_num, header_title in enumerate(all_headers, 1):
                cell = ws.cell(row=2, column=col_num)
                cell.value = header_title
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                
                # Atur lebar kolom otomatis sedikit
                if col_num == 2: # Kolom NIS
                    ws.column_dimensions[get_column_letter(col_num)].width = 15
                elif col_num == 3: # Kolom Nama
                    ws.column_dimensions[get_column_letter(col_num)].width = 30
                else: # Kolom tanggal & total
                    ws.column_dimensions[get_column_letter(col_num)].width = 5

            # --- Data Processing ---
            siswas = Siswa.original_objects.filter(kelas=kelas).order_by('fullname')
            absensi_qs = Absensi.original_objects.filter(
                siswa__kelas=kelas,
                date__year=year,
                date__month=month
            )

            # Mapping data absensi agar mudah diakses: {(siswa_id, day): status}
            absensi_map = {}
            for a in absensi_qs:
                absensi_map[(a.siswa_id, a.date.day)] = a.status

            # --- Tulis Data Siswa ---
            row_num = 3
            for idx, siswa in enumerate(siswas, 1):
                # Kolom Identitas
                ws.cell(row=row_num, column=1).value = idx
                ws.cell(row=row_num, column=2).value = siswa.nis
                ws.cell(row=row_num, column=3).value = siswa.fullname

                # Inisialisasi Counter
                counts = {'hadir': 0, 'sakit': 0, 'izin': 0, 'alfa': 0, 'bolos': 0}

                # Kolom Tanggal (Loop 1 sampai tanggal terakhir bulan itu)
                col_offset = 4 # Kolom 1,2,3 sudah terpakai
                for day in range(1, num_days + 1):
                    status = absensi_map.get((siswa.id, day))
                    cell = ws.cell(row=row_num, column=col_offset + day - 1)
                    
                    if status:
                        # Ambil huruf depan status (H, S, I, A, B) kapital
                        short_status = status[0].upper() 
                        cell.value = short_status
                        cell.alignment = center_align
                        
                        # Hitung total
                        if status in counts:
                            counts[status] += 1
                    else:
                        cell.value = "-" # Tidak ada data/libur
                        cell.alignment = center_align

                # Kolom Total (Summary)
                summary_start_col = col_offset + num_days
                ws.cell(row=row_num, column=summary_start_col).value = counts['hadir']
                ws.cell(row=row_num, column=summary_start_col + 1).value = counts['sakit']
                ws.cell(row=row_num, column=summary_start_col + 2).value = counts['izin']
                ws.cell(row=row_num, column=summary_start_col + 3).value = counts['alfa']
                ws.cell(row=row_num, column=summary_start_col + 4).value = counts['bolos']

                row_num += 1

            wb.save(response)
            return response

        except Exception as e:
            return HttpResponse(f"Terjadi kesalahan: {str(e)}")